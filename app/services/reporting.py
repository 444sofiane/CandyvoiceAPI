"""Construit le rapport admin d'usage/satisfaction et l'envoie par e-mail
via SMTP2GO. Lit directement dans les collections Firestore
usageEvents/satisfactionEvents (la même source depuis laquelle ces
données sont synchronisées *vers* Zoho CRM), donc le rapport est toujours
au moins aussi frais que tout ce qu'il y a dans Zoho Analytics et n'a pas
besoin d'identifiants API Zoho Analytics séparés.
"""
import io
import smtplib
from datetime import datetime, timedelta, timezone
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app import config
from app.services.firebase import get_firestore_client

# Libellés plus lisibles pour le classeur que les clés internes brutes des
# fonctionnalités. "noizoff" est une ancienne donnée de test enregistrée
# avant le correctif de cohérence de nommage (auth-guard.js envoie
# maintenant "noiseFilter" pour correspondre à config.FEATURE_KEY_*) —
# gardé comme libellé historique distinct plutôt que fusionné dans
# "NoizOff", pour que les anciennes lignes ne soient pas silencieusement
# comptées avec les nouvelles sous un même nom.
FEATURE_DISPLAY_NAMES = {
    "noiseFilter": "NoizOff",
    "imitation": "Imitation vocale",
    "deepfake": "Détection de deepfake",
    "frameRecovery": "Récupération de trames",
    "unknown": "Non spécifié",
}

# Anciennes/autres orthographes qui doivent être comptées comme la même
# fonctionnalité pendant l'agrégation. "noizoff" est ce qu'envoyait le lien
# du sondage avant d'être aligné pour correspondre au "noiseFilter" de
# Usage_Events — même fonctionnalité, juste enregistrée sous une chaîne
# différente pendant un temps. Appliqué à la lecture dans build_report(),
# donc ça fusionne proprement en une seule ligne "NoizOff" sans toucher
# aux enregistrements Zoho/Firestore sous-jacents.
FEATURE_ALIASES = {
    "noizoff": "noiseFilter",
}


def _normalize_feature(feature):
    feature = feature or "unknown"
    return FEATURE_ALIASES.get(feature, feature)


_HEADER_FILL = PatternFill(start_color="5A4BFF", end_color="5A4BFF", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)

RANGE_LABELS = {
    "day": "Dernier jour",
    "week": "7 derniers jours",
    "3months": "3 derniers mois",
    "6months": "6 derniers mois",
}

# Les périodes plus longues sont regroupées par mois plutôt que par jour,
# pour qu'un e-mail sur 6 mois reste une poignée de lignes plutôt que
# ~180 — les périodes jour/semaine sont assez courtes pour qu'une
# répartition par jour reste lisible.
_MONTH_BUCKETED_RANGES = {"3months", "6months"}

_RANGE_TO_DAYS = {
    "day": 1,
    "week": 7,
    # Les mois calendaires ont une longueur variable ; un nombre de jours
    # fixe est une approximation suffisante pour une fenêtre de rapport
    # (pas utilisé pour quoi que ce soit de sensible côté facturation ou
    # légal). Évite d'ajouter une dépendance dateutil juste pour ça.
    "3months": 91,
    "6months": 182,
}


def resolve_date_range(range_key, reference_date=None):
    """Retourne (start, end) comme datetimes UTC, end étant la fin de
    reference_date (par défaut aujourd'hui) et start étant `range_key` en
    arrière depuis là, inclus."""
    if range_key not in _RANGE_TO_DAYS:
        raise ValueError(f"Unknown range: {range_key!r}. Expected one of {sorted(_RANGE_TO_DAYS)}.")

    if reference_date is None:
        reference_date = datetime.now(timezone.utc).date()

    end = datetime.combine(reference_date, datetime.max.time(), tzinfo=timezone.utc)
    start = end - timedelta(days=_RANGE_TO_DAYS[range_key]) + timedelta(microseconds=1)
    return start, end


def _bucket_label(dt, range_key):
    if range_key in _MONTH_BUCKETED_RANGES:
        return dt.strftime("%Y-%m")
    return dt.strftime("%Y-%m-%d")


def _fetch_events(collection_name, start, end):
    client = get_firestore_client()
    query = (
        client.collection(collection_name)
        .where("createdAt", ">=", start)
        .where("createdAt", "<=", end)
    )
    docs = []
    for snap in query.stream():
        data = snap.to_dict() or {}
        created_at = data.get("createdAt")
        # Les timestamps Firestore reviennent déjà comme des objets datetime.
        if created_at is not None:
            docs.append(data)
    return docs


def build_report(range_key, reference_date=None):
    """Retourne un dict avec tout ce dont render_report_html a besoin : la
    fenêtre de dates résolue, une table d'usage regroupée (utilisateurs
    uniques + nombre d'événements par bucket par fonctionnalité), une
    table de satisfaction regroupée (score moyen + nombre de réponses par
    bucket par fonctionnalité), et les totaux sur toute la période par
    fonctionnalité pour un résumé rapide en tête."""
    start, end = resolve_date_range(range_key, reference_date)

    usage_events = _fetch_events("usageEvents", start, end)
    satisfaction_events = _fetch_events("satisfactionEvents", start, end)

    # bucket -> feature -> set(uid)   (utilisateurs uniques, dédupliqués par bucket)
    usage_buckets = {}
    # feature -> set(uid) sur toute la période, pour le résumé en tête
    usage_totals = {}

    for event in usage_events:
        feature = _normalize_feature(event.get("feature"))
        uid = event.get("uid")
        created_at = event.get("createdAt")
        if created_at is None or uid is None:
            continue
        bucket = _bucket_label(created_at, range_key)
        usage_buckets.setdefault(bucket, {}).setdefault(feature, set()).add(uid)
        usage_totals.setdefault(feature, set()).add(uid)

    # bucket -> feature -> list[score]
    satisfaction_buckets = {}
    # feature -> list[score] sur toute la période
    satisfaction_totals = {}

    for event in satisfaction_events:
        feature = _normalize_feature(event.get("feature"))
        score = event.get("score")
        created_at = event.get("createdAt")
        if created_at is None or score is None:
            continue
        try:
            score = float(score)
        except (TypeError, ValueError):
            continue
        bucket = _bucket_label(created_at, range_key)
        satisfaction_buckets.setdefault(bucket, {}).setdefault(feature, []).append(score)
        satisfaction_totals.setdefault(feature, []).append(score)

    return {
        "range_key": range_key,
        "range_label": RANGE_LABELS[range_key],
        "start": start,
        "end": end,
        "usage_buckets": usage_buckets,
        "usage_totals": {feature: len(uids) for feature, uids in usage_totals.items()},
        "satisfaction_buckets": satisfaction_buckets,
        "satisfaction_totals": {
            feature: {"average": sum(scores) / len(scores), "count": len(scores)}
            for feature, scores in satisfaction_totals.items()
        },
    }


def _fmt_avg(value):
    return f"{value:.1f}"


def _display_feature(feature):
    return FEATURE_DISPLAY_NAMES.get(feature, feature)


def _style_header_row(ws, row_idx, num_cols):
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autofit_columns(ws, widths):
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_summary_sheet(ws, report):
    features = sorted(set(report["usage_totals"]) | set(report["satisfaction_totals"]))

    ws["A1"] = "Rapport d'utilisation et de satisfaction CandyVoice"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = report["range_label"]
    ws["A3"] = f"{report['start'].date()} — {report['end'].date()}"
    ws["A2"].font = ws["A3"].font = Font(color="666666")

    header_row = 5
    headers = ["Fonctionnalité", "Utilisateurs uniques", "Satisfaction moyenne", "Évaluations"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=header)
    _style_header_row(ws, header_row, len(headers))

    for offset, feature in enumerate(features, start=1):
        row = header_row + offset
        satisfaction = report["satisfaction_totals"].get(feature)
        ws.cell(row=row, column=1, value=_display_feature(feature))
        ws.cell(row=row, column=2, value=report["usage_totals"].get(feature, 0))
        ws.cell(row=row, column=3, value=round(satisfaction["average"], 1) if satisfaction else None)
        ws.cell(row=row, column=4, value=satisfaction["count"] if satisfaction else 0)

    last_row = header_row + len(features)
    _autofit_columns(ws, [26, 14, 16, 10])

    if not features:
        ws.cell(row=header_row + 1, column=1, value="Aucune activité sur cette période.")
        return

    # Utilisateurs uniques par fonctionnalité
    users_chart = BarChart()
    users_chart.title = "Utilisateurs uniques par fonctionnalité"
    users_chart.y_axis.title = "Utilisateurs uniques"
    users_data = Reference(ws, min_col=2, min_row=header_row, max_row=last_row)
    categories = Reference(ws, min_col=1, min_row=header_row + 1, max_row=last_row)
    users_chart.add_data(users_data, titles_from_data=True)
    users_chart.set_categories(categories)
    users_chart.width, users_chart.height = 14, 8
    ws.add_chart(users_chart, f"F{header_row}")

    # Satisfaction moyenne par fonctionnalité
    satisfaction_chart = BarChart()
    satisfaction_chart.title = "Satisfaction moyenne par fonctionnalité (0-10)"
    satisfaction_chart.y_axis.title = "Score moyen"
    satisfaction_chart.y_axis.scaling.min = 0
    satisfaction_chart.y_axis.scaling.max = 10
    satisfaction_data = Reference(ws, min_col=3, min_row=header_row, max_row=last_row)
    satisfaction_chart.add_data(satisfaction_data, titles_from_data=True)
    satisfaction_chart.set_categories(categories)
    satisfaction_chart.width, satisfaction_chart.height = 14, 8
    ws.add_chart(satisfaction_chart, f"F{header_row + 17}")


def _pivot_usage_by_period(report):
    periods = sorted(report["usage_buckets"].keys())
    features = sorted({f for bucket in report["usage_buckets"].values() for f in bucket})
    rows = [
        [len(report["usage_buckets"].get(period, {}).get(feature, set())) for feature in features]
        for period in periods
    ]
    return periods, features, rows


def _pivot_satisfaction_by_period(report):
    periods = sorted(report["satisfaction_buckets"].keys())
    features = sorted({f for bucket in report["satisfaction_buckets"].values() for f in bucket})
    rows = []
    for period in periods:
        row = []
        for feature in features:
            scores = report["satisfaction_buckets"].get(period, {}).get(feature, [])
            row.append(round(sum(scores) / len(scores), 1) if scores else None)
        rows.append(row)
    return periods, features, rows


def _write_pivot_sheet(ws, periods, features, rows, value_label, chart_cls, bucket_word):
    ws["A1"] = f"{value_label} par {bucket_word} et fonctionnalité"
    ws["A1"].font = Font(bold=True, size=12)

    header_row = 3
    ws.cell(row=header_row, column=1, value=bucket_word.capitalize())
    for col_idx, feature in enumerate(features, start=2):
        ws.cell(row=header_row, column=col_idx, value=_display_feature(feature))
    _style_header_row(ws, header_row, len(features) + 1)

    for row_offset, period in enumerate(periods, start=1):
        row = header_row + row_offset
        ws.cell(row=row, column=1, value=period)
        for col_idx, value in enumerate(rows[row_offset - 1], start=2):
            ws.cell(row=row, column=col_idx, value=value)

    _autofit_columns(ws, [14] + [22] * len(features))

    if not periods or not features:
        ws.cell(row=header_row + 1, column=1, value="Aucune activité sur cette période.")
        return

    last_row = header_row + len(periods)
    chart = chart_cls()
    chart.title = f"{value_label} dans le temps"
    data = Reference(ws, min_col=2, max_col=1 + len(features), min_row=header_row, max_row=last_row)
    categories = Reference(ws, min_col=1, min_row=header_row + 1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.width, chart.height = 20, 10
    ws.add_chart(chart, f"A{last_row + 3}")


def build_report_workbook(report):
    """Construit la pièce jointe .xlsx : une feuille Résumé (totaux par
    fonctionnalité + deux graphiques en barres) et deux feuilles de détail
    pivotées période-par-fonctionnalité (usage, satisfaction), chacune
    avec son propre graphique de tendance. Utiliser un vrai classeur
    plutôt qu'un CSV évite deux problèmes à la fois : les CSV délimités
    par virgule s'ouvrent tout brouillés dans Excel sous les locales qui
    attendent ";" comme séparateur de liste (ex. le français), et un CSV
    ne peut pas du tout contenir de graphiques — c'est du texte brut.
    """
    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Résumé"
    _write_summary_sheet(summary_ws, report)

    bucket_word = "mois" if report["range_key"] in _MONTH_BUCKETED_RANGES else "jour"

    usage_ws = wb.create_sheet("Détail utilisation")
    _write_pivot_sheet(
        usage_ws, *_pivot_usage_by_period(report),
        value_label="Utilisateurs uniques", chart_cls=BarChart, bucket_word=bucket_word,
    )

    satisfaction_ws = wb.create_sheet("Détail satisfaction")
    _write_pivot_sheet(
        satisfaction_ws, *_pivot_satisfaction_by_period(report),
        value_label="Satisfaction moyenne", chart_cls=LineChart, bucket_word=bucket_word,
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def render_report_html(report):
    features = sorted(
        set(report["usage_totals"]) | set(report["satisfaction_totals"])
    )

    summary_rows = "".join(
        f"<tr>"
        f"<td style='padding:6px 12px;border-bottom:1px solid #eee'>{feature}</td>"
        f"<td style='padding:6px 12px;border-bottom:1px solid #eee;text-align:right'>"
        f"{report['usage_totals'].get(feature, 0)}</td>"
        f"<td style='padding:6px 12px;border-bottom:1px solid #eee;text-align:right'>"
        f"{_fmt_avg(report['satisfaction_totals'][feature]['average']) if feature in report['satisfaction_totals'] else '—'}</td>"
        f"<td style='padding:6px 12px;border-bottom:1px solid #eee;text-align:right'>"
        f"{report['satisfaction_totals'][feature]['count'] if feature in report['satisfaction_totals'] else 0}</td>"
        f"</tr>"
        for feature in features
    ) or "<tr><td colspan='4' style='padding:12px'>Aucune activité sur cette période.</td></tr>"

    bucket_labels = sorted(set(report["usage_buckets"]) | set(report["satisfaction_buckets"]))
    detail_rows = ""
    for bucket in bucket_labels:
        usage_by_feature = report["usage_buckets"].get(bucket, {})
        satisfaction_by_feature = report["satisfaction_buckets"].get(bucket, {})
        bucket_features = sorted(set(usage_by_feature) | set(satisfaction_by_feature))
        for feature in bucket_features:
            unique_users = len(usage_by_feature.get(feature, set()))
            scores = satisfaction_by_feature.get(feature, [])
            avg_score = _fmt_avg(sum(scores) / len(scores)) if scores else "—"
            detail_rows += (
                f"<tr>"
                f"<td style='padding:4px 12px;border-bottom:1px solid #f2f2f2'>{bucket}</td>"
                f"<td style='padding:4px 12px;border-bottom:1px solid #f2f2f2'>{feature}</td>"
                f"<td style='padding:4px 12px;border-bottom:1px solid #f2f2f2;text-align:right'>{unique_users}</td>"
                f"<td style='padding:4px 12px;border-bottom:1px solid #f2f2f2;text-align:right'>{avg_score}</td>"
                f"</tr>"
            )
    if not detail_rows:
        detail_rows = "<tr><td colspan='4' style='padding:12px'>Aucune activité sur cette période.</td></tr>"

    period = f"{report['start'].date()} — {report['end'].date()}"

    return f"""
    <html><body style="font-family:Arial,Helvetica,sans-serif;color:#222">
      <h2 style="margin-bottom:0">Rapport d'utilisation et de satisfaction CandyVoice</h2>
      <p style="color:#666;margin-top:4px">{report['range_label']} ({period})</p>

      <h3>Résumé</h3>
      <table style="border-collapse:collapse;width:100%;max-width:640px">
        <thead>
          <tr style="text-align:left;border-bottom:2px solid #ccc">
            <th style="padding:6px 12px">Fonctionnalité</th>
            <th style="padding:6px 12px;text-align:right">Utilisateurs uniques</th>
            <th style="padding:6px 12px;text-align:right">Satisfaction moyenne</th>
            <th style="padding:6px 12px;text-align:right">Évaluations</th>
          </tr>
        </thead>
        <tbody>{summary_rows}</tbody>
      </table>

      <h3 style="margin-top:32px">Répartition par {'mois' if report['range_key'] in {'3months', '6months'} else 'jour'}</h3>
      <table style="border-collapse:collapse;width:100%;max-width:640px;font-size:13px">
        <thead>
          <tr style="text-align:left;border-bottom:2px solid #ccc">
            <th style="padding:4px 12px">Période</th>
            <th style="padding:4px 12px">Fonctionnalité</th>
            <th style="padding:4px 12px;text-align:right">Utilisateurs uniques</th>
            <th style="padding:4px 12px;text-align:right">Satisfaction moyenne</th>
          </tr>
        </thead>
        <tbody>{detail_rows}</tbody>
      </table>

      <p style="color:#999;font-size:12px;margin-top:32px">
        Généré automatiquement à partir des données usageEvents / satisfactionEvents de CandyVoice.
      </p>
    </body></html>
    """


def send_report_email(report):
    if not config.SMTP2GO_USERNAME or not config.SMTP2GO_PASSWORD:
        raise RuntimeError(
            "SMTP2GO_USERNAME / SMTP2GO_PASSWORD are not configured — set them as "
            "environment variables before sending admin reports."
        )
    if not config.ADMIN_REPORT_RECIPIENTS:
        raise RuntimeError("ADMIN_REPORT_RECIPIENTS is empty — nothing to send the report to.")

    period = f"{report['start'].date()} — {report['end'].date()}"
    workbook_bytes = build_report_workbook(report)
    filename = f"candyvoice-report-{report['range_key']}-{report['end'].date()}.xlsx"

    message = MIMEMultipart("mixed")
    message["Subject"] = f"Rapport CandyVoice — {report['range_label']}"
    message["From"] = f"{config.SMTP2GO_FROM_NAME} <{config.SMTP2GO_FROM_EMAIL}>"
    message["To"] = ", ".join(config.ADMIN_REPORT_RECIPIENTS)

    body_text = (
        f"Rapport d'utilisation et de satisfaction CandyVoice — {report['range_label']} ({period}).\n\n"
        f"Consultez le classeur joint ({filename}) : une feuille Résumé avec les totaux "
        f"et graphiques par fonctionnalité, ainsi que des feuilles Détail utilisation / "
        f"Détail satisfaction réparties par "
        f"{'mois' if report['range_key'] in _MONTH_BUCKETED_RANGES else 'jour'}, "
        f"chacune avec son propre graphique d'évolution."
    )
    message.attach(MIMEText(body_text, "plain"))

    attachment = MIMEApplication(
        workbook_bytes,
        _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    attachment.add_header("Content-Disposition", "attachment", filename=filename)
    message.attach(attachment)

    with smtplib.SMTP(config.SMTP2GO_HOST, config.SMTP2GO_PORT) as server:
        server.starttls()
        server.login(config.SMTP2GO_USERNAME, config.SMTP2GO_PASSWORD)
        server.sendmail(
            config.SMTP2GO_FROM_EMAIL,
            config.ADMIN_REPORT_RECIPIENTS,
            message.as_string(),
        )
