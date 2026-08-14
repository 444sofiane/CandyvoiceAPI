"""Rapport mensuel d'usage des clés API, agrégé par utilisateur, envoyé
par e-mail aux admins. Complète le rapport d'usage/satisfaction du site de
reporting.py — module séparé parce que les données source sont
complètement différentes (apiKeys / apiKeyUsage, pas
usageEvents/satisfactionEvents) même si la forme "agréger quelque chose,
construire un classeur, l'envoyer par e-mail via SMTP2GO" suit le même
schéma.

Les quatre fonctionnalités sont comptées en secondes de traitement, pas en
nombre de fichiers, puisque l'enforcement par clé API est par budget de
session — voir api_key_quota.py.
"""
import io
import smtplib
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app import config
from app.services.api_key_quota import ALL_FEATURE_KEYS, current_period, usage_doc_ref
from app.services.api_keys import list_all_key_records, resolve_emails
from app.services.firebase import get_firestore_client
from app.services.reporting import FEATURE_DISPLAY_NAMES

_HEADER_FILL = PatternFill(start_color="5A4BFF", end_color="5A4BFF", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def build_api_usage_report(period: str | None = None) -> dict:
    """Agrège l'usage de chaque clé API pour `period` ("YYYY-MM", par
    défaut le mois calendaire UTC en cours) en une ligne par utilisateur —
    les mêmes compteurs par clé/par fonctionnalité sous-jacents que GET
    /api/keys/{key_id}/usage expose, juste sommés sur toutes les clés d'un
    utilisateur plutôt que montrés une clé à la fois. Toutes les
    fonctionnalités sont sommées en secondes traitées (compteur purement
    indicatif — voir api_key_quota.log_session_usage) ; le budget qui
    bloque réellement une requête est appliqué par session, pas récupéré
    depuis ce total mensuel.

    Inclut l'usage des clés révoquées — une clé révoquée en milieu de mois
    a quand même consommé du budget ce mois-là, et ceci est un rapport
    d'usage, pas une liste de clés actives. Les utilisateurs sans aucune
    activité sur cette période sont omis, pour que le rapport reste centré
    sur l'activité réelle plutôt que de lister chaque utilisateur inscrit.

    Retourne {"period": ..., "rows": [{"uid", "email", "plans",
    "key_count", "usage_seconds": {feature_key: float}, "total_seconds":
    float}, ...]}, lignes triées par usage total décroissant."""
    period = period or current_period()
    firestore_client = get_firestore_client()
    key_records = list_all_key_records()

    per_user = {}
    for record in key_records:
        uid = record.get("uid")
        if not uid:
            continue

        entry = per_user.setdefault(uid, {
            "plans": set(),
            "key_count": 0,
            "usage_seconds": {feature_key: 0.0 for feature_key in ALL_FEATURE_KEYS},
        })
        entry["plans"].add(record["plan"])
        entry["key_count"] += 1

        snapshot = usage_doc_ref(firestore_client, record["key_id"], period).get()
        data = snapshot.to_dict() if snapshot.exists else {}
        for feature_key in ALL_FEATURE_KEYS:
            entry["usage_seconds"][feature_key] += (data.get(feature_key) or {}).get("secondsUsed", 0)

    emails = resolve_emails(list(per_user.keys()))

    rows = []
    for uid, entry in per_user.items():
        total_seconds = sum(entry["usage_seconds"].values())
        if total_seconds == 0:
            continue
        rows.append({
            "uid": uid,
            "email": emails.get(uid),
            "plans": sorted(entry["plans"]),
            "key_count": entry["key_count"],
            "usage_seconds": entry["usage_seconds"],
            "total_seconds": total_seconds,
        })
    rows.sort(key=lambda row: row["total_seconds"], reverse=True)

    return {"period": period, "rows": rows}


def _style_header_row(ws, row_idx, num_cols):
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def build_api_usage_report_workbook(report: dict) -> bytes:
    """Une feuille, une ligne par utilisateur avec de l'usage cette
    période : e-mail, uid, offre(s), nombre de clés, une colonne de
    secondes traitées par fonctionnalité, et un total en secondes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Utilisation par utilisateur"

    ws["A1"] = "Rapport d'utilisation de l'API CandyVoice"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = report["period"]
    ws["A2"].font = Font(color="666666")

    header_row = 4
    headers = (
        ["E-mail", "Uid", "Offre(s)", "Clés"]
        + [f"{FEATURE_DISPLAY_NAMES.get(fk, fk)} (s)" for fk in ALL_FEATURE_KEYS]
        + ["Total (secondes)"]
    )
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=header)
    _style_header_row(ws, header_row, len(headers))

    total_col = 4 + len(ALL_FEATURE_KEYS) + 1

    for offset, row in enumerate(report["rows"], start=1):
        r = header_row + offset
        ws.cell(row=r, column=1, value=row["email"] or "(aucun e-mail enregistré)")
        ws.cell(row=r, column=2, value=row["uid"])
        ws.cell(row=r, column=3, value=", ".join(row["plans"]))
        ws.cell(row=r, column=4, value=row["key_count"])
        for col_offset, feature_key in enumerate(ALL_FEATURE_KEYS, start=5):
            ws.cell(row=r, column=col_offset, value=round(row["usage_seconds"][feature_key]))
        ws.cell(row=r, column=total_col, value=round(row["total_seconds"]))

    if not report["rows"]:
        ws.cell(row=header_row + 1, column=1, value="Aucune utilisation de clé API sur cette période.")

    _autofit = [28, 24, 16, 8] + [16] * len(ALL_FEATURE_KEYS) + [16]
    for col_idx, width in enumerate(_autofit, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def send_api_usage_report_email(report: dict) -> None:
    if not config.SMTP2GO_USERNAME or not config.SMTP2GO_PASSWORD:
        raise RuntimeError(
            "SMTP2GO_USERNAME / SMTP2GO_PASSWORD are not configured — set them as "
            "environment variables before sending admin reports."
        )
    if not config.ADMIN_REPORT_RECIPIENTS:
        raise RuntimeError("ADMIN_REPORT_RECIPIENTS is empty — nothing to send the report to.")

    workbook_bytes = build_api_usage_report_workbook(report)
    filename = f"candyvoice-api-usage-{report['period']}.xlsx"

    message = MIMEMultipart("mixed")
    message["Subject"] = f"Rapport d'utilisation de l'API CandyVoice — {report['period']}"
    message["From"] = f"{config.SMTP2GO_FROM_NAME} <{config.SMTP2GO_FROM_EMAIL}>"
    message["To"] = ", ".join(config.ADMIN_REPORT_RECIPIENTS)

    total_users = len(report["rows"])
    total_seconds = round(sum(row["total_seconds"] for row in report["rows"]))
    body_text = (
        f"Rapport d'utilisation des clés API CandyVoice — {report['period']}.\n\n"
        f"{total_users} utilisateur(s) avec une activité de clé API sur cette période, "
        f"{total_seconds} seconde(s) d'audio traitées au total, toutes clés et "
        f"fonctionnalités confondues.\n\n"
        f"Consultez le classeur joint ({filename}) pour le détail par utilisateur et par fonctionnalité."
    )
    message.attach(MIMEText(body_text, "plain"))

    attachment = MIMEApplication(
        workbook_bytes,
        _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    attachment.add_header("Content-Disposition", "attachment", filename=filename)
    message.attach(attachment)

    with smtplib.SMTP(config.SMTP2GO_HOST, config.SMTP2GO_PORT) as server:
        server.starttls()
        server.login(config.SMTP2GO_USERNAME, config.SMTP2GO_PASSWORD)
        server.sendmail(
            config.SMTP2GO_FROM_EMAIL,
            config.ADMIN_REPORT_RECIPIENTS,
            message.as_string(),
        )
